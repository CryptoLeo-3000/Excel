import openpyxl as xl

def insert_cell(filename):
    wb = xl.load_workbook(f'Excel/{filename}.xlsx')
    sheet = wb['Sheet1']
    cell1 = sheet['A1']

    for row in range(2, sheet.max_row + 1):
        ce = sheet.cell(row, 3)
        correction = ce.value * 0.9
        correction_cell = sheet.cell(row, 4)
        correction_cell.value = correction

    wb.save(f'Excel/{filename}.xlsx')